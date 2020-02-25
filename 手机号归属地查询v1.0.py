from requests_html import HTMLSession
import pandas as pd
import time
import random
from openpyxl import load_workbook
import os
s = HTMLSession()


header={"User-Agent":"Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36"}

def PhoneNumber_search(phone_list):     # 手机号查询
    """
    phone_list:     待查询的手机号列表

    """
    for phone in phone_list:
        try:
            r = s.get(url="http://www.ip138.com:8080/search.asp?mobile="+str(phone)+"&action=mobile",headers=header)
            r.encoding=r.apparent_encoding  # 设置编码
            Card_number_home = r.html.xpath("/html/body/table[2]/tr[3]/td[2]/text()")[0] # 卡号归属地
            Card_type = r.html.xpath("/html/body/table[2]/tr[4]/td[2]/text()")[0] # 卡类型
            Area_code = r.html.xpath("/html/body/table[2]/tr[5]/td[2]/text()")[0] # 区号
            post_code = r.html.xpath("/html/body/table[2]/tr[6]/td[2]/text()")[0] # 邮编
            print("您查询的手机号码(段)为：",phone)
            print("卡号归属地:",Card_number_home)
            print("卡类型:",Card_type)
            print("区号:",Area_code)
            print("邮编:",post_code)
            print("-"*100)
            time.sleep(random.uniform(0,1))   # 设置随机休眠时间，以防止短时间内被封IP。
        except:
            print("查询出错，请检查填写的手机号(段)或本机网络状态是否正常可用！")
            continue
        

def PhoneNumber_search_batch(file_read_path,file_save_path):     # 手机号查询,可批量操作(从.xlsx文件中读取手机号)
    """
    file_read_path:     读取文件的路径

    file_save_path      保存文件的路径
    """
    phone_table = pd.read_excel(io=file_read_path)   # 获取Excel文件内容并将数据转为DataFrame类型

    for i in range(phone_table.shape[0]):   # 遍历每行数据
        try:
            phone=phone_table.loc[i,'手机号']   # 提取手机号
            r = s.get(url="http://www.ip138.com:8080/search.asp?mobile="+str(phone)+"&action=mobile",headers=header)
            r.encoding=r.apparent_encoding  # 设置编码
            Card_number_home = r.html.xpath("/html/body/table[2]/tr[3]/td[2]/text()")[0] # 卡号归属地
            Card_type = r.html.xpath("/html/body/table[2]/tr[4]/td[2]/text()")[0] # 卡类型
            Area_code = r.html.xpath("/html/body/table[2]/tr[5]/td[2]/text()")[0] # 区号
            post_code = r.html.xpath("/html/body/table[2]/tr[6]/td[2]/text()")[0] # 邮编

            phone_table.loc[i,'卡号归属地']=Card_number_home
            phone_table.loc[i,'卡类型']=Card_type
            phone_table.loc[i,'区号']=str(Area_code)
            phone_table.loc[i,'邮编']=str(post_code)

            print(phone_table.loc[i])
            print("-"*100)
            time.sleep(random.uniform(0,1))   # 设置随机休眠时间，以防止短时间内被封IP。
        except:
            print("查询出错，请检查文件中的手机号(段)或本机网络状态是否正常可用！")
            continue
    
    # 方法一： 使用此方法保存的表会不会覆盖工作簿中已存在的其他表，保存形式为在工作簿中追加一张或多张表并对应的写入数据。
    # with pd.ExcelWriter(file_read_path,engine='openpyxl') as writer:       # 使用pd.ExcelWriter建立一个writer,使用with关键字不需要再单独写save和close自动完成。
    #     writer.book = load_workbook(file_read_path)
    #     phone_table.to_excel(writer,sheet_name="手机号(段)查询结果",index = False)     # 设置index = False，保存后的表中没有行索引号。
    # print(writer.book.sheetnames)

    # 方法二：使用此方法保存的表会不会覆盖工作簿中已存在的其他表，保存形式为在工作簿中追加一张或多张表并对应的写入数据。
    # wb = load_workbook(file_read_path)
    # writer = pd.ExcelWriter(file_read_path,engine='openpyxl')
    # writer.book = wb
    # phone_table.to_excel(writer,sheet_name="手机号(段)查询结果",index = False)
    # wb.save(file_read_path)
    # wb.close()
    # print(wb.sheetnames)
    
    # 方法三：使用此方法保存的表会覆盖工作簿中已存在的所有表，新工作簿可使用此方法一次保存一张或多张表。
    with pd.ExcelWriter(file_save_path) as writer:                   # 使用pd.ExcelWriter建立一个writer,使用with关键字不需要再单独写save和close自动完成。     
        phone_table.to_excel(writer,sheet_name="手机号(段)查询结果",index = False)    # 设置index = False，保存后的表中没有行索引号。


if __name__ == "__main__":
    file_read_path = "C:\\Users\\Administrator\\Desktop\\phone.xlsx"             # 读取文件的路径
    file_save_path = "C:\\Users\\Administrator\\Desktop\\手机号(段)查询结果.xlsx"     # 保存文件的路径
    # phone_list = [14752996388]     # 请在中括号内填写需要查询的手机号(段)，以英文逗号隔开。
    # PhoneNumber_search(phone_list)  # 手机号(段)查询
    PhoneNumber_search_batch(file_read_path,file_save_path) # 手机号(段)查询,可批量操作(从.xlsx文件中读取手机号)


